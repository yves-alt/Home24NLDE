import csv
import io
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

from database.database import get_connection


CATEGORY_KEYWORDS = {
    "kitchen": ["küche", "keuken", "herd", "spüle", "schrank"],
    "bathroom": ["bad", "dusche", "wanne", "waschbecken", "badezimmer"],
    "bedroom": ["bett", "schlaf", "matratze", "kissen", "decke"],
    "living": ["sofa", "couch", "wohnzimmer", "fernseher", "tisch"],
    "outdoor": ["garten", "terrasse", "outdoor", "balkon"],
    "textile": ["stoff", "textil", "kissen", "decke", "vorhang"],
    "lighting": ["lampe", "leuchte", "licht", "led", "pendel"],
    "storage": ["schrank", "regal", "kommode", "schublade"],
    "color": ["farb", "farbig", "bunt", "weiss", "schwarz", "blau"],
    "material": ["holz", "metall", "stoff", "leder", "glas", "eiche"],
}


def normalize_segment(text: str) -> str:
    if not text:
        return ""
    text = text.lower().strip()
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"\s+", " ", text)
    return text


def detect_category(source: str, target: str) -> str:
    combined = (source + " " + target).lower()
    for cat, keywords in CATEGORY_KEYWORDS.items():
        if any(kw in combined for kw in keywords):
            return cat
    return "general"


# ── Public entry points ────────────────────────────────────────────────

def import_tm_from_bytes(file_bytes: bytes, filename: str, progress_callback=None) -> dict:
    """Parse a TM file from raw bytes (xlsx or csv) and UPSERT into the DB.

    Never deletes existing TM data. Returns stats dict with keys:
    inserted, updated, duplicates, invalid, total.
    """
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xls"):
        parsed = _parse_excel_bytes(file_bytes)
    elif suffix == ".csv":
        parsed = _parse_csv_bytes(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Accepted: .xlsx, .csv")
    return _upsert_rows(parsed, progress_callback)


def import_tm_from_excel(filepath: str, progress_callback=None) -> dict:
    """Legacy file-path-based import. Delegates to import_tm_from_bytes."""
    data = Path(filepath).read_bytes()
    return import_tm_from_bytes(data, Path(filepath).name, progress_callback)


# ── Parsers ────────────────────────────────────────────────────────────

def _parse_excel_bytes(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Prefer sheets whose name hints at translation units
    sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if any(k in nl for k in ("translation unit", " tu", "tm", "memory", "segment")):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(c).lower().strip() if c else "" for c in rows[0]]
    source_col = _find_col(header, ["source", "de-de", "source (de-de)", "deutsch", "german", "source_segment"])
    target_col = _find_col(header, ["target", "nl-nl", "target (nl-nl)", "dutch", "dutch (nl-nl)", "target_segment", "nl"])
    freq_col   = _find_col(header, ["usage count", "usage", "frequency", "freq"])
    date_col   = _find_col(header, ["creation date", "created", "created_at", "date"])
    by_col     = _find_col(header, ["created by", "author", "created_by"])
    id_col     = _find_col(header, ["id"])

    if source_col is None or target_col is None:
        raise ValueError(
            f"Cannot find source/target columns. "
            f"Detected headers: {header}. "
            f"Expected columns named: source/de-de and target/nl-nl (or deutsch/dutch)."
        )

    result = []
    for row in rows[1:]:
        source = _cell_str(row, source_col)
        target = _cell_str(row, target_col)
        if not source or not target:
            result.append(None)
            continue
        result.append({
            "source": source,
            "target": target,
            "freq": _cell_int(row, freq_col),
            "created": _cell_str(row, date_col) or None,
            "created_by": _cell_str(row, by_col) or None,
            "source_id": _cell_int(row, id_col) or None,
        })
    return result


def _parse_csv_bytes(file_bytes: bytes) -> list:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Cannot decode CSV. Expected UTF-8 or Latin-1 encoding.")

    reader = csv.DictReader(io.StringIO(text))
    raw_fields = reader.fieldnames or []
    fieldnames = [f.lower().strip() for f in raw_fields]

    source_key = _find_key(fieldnames, ["source", "de-de", "german", "source_segment", "deutsch"])
    target_key = _find_key(fieldnames, ["target", "nl-nl", "dutch", "target_segment", "nl"])

    if not source_key or not target_key:
        raise ValueError(
            f"Cannot find source/target columns in CSV. "
            f"Headers found: {fieldnames}. "
            f"Expected columns named: source/de-de and target/nl-nl."
        )

    result = []
    for row in reader:
        # Normalize keys
        norm_row = {k.lower().strip(): v for k, v in row.items()}
        source = (norm_row.get(source_key) or "").strip()
        target = (norm_row.get(target_key) or "").strip()
        if not source or not target:
            result.append(None)
            continue
        freq_raw = (
            norm_row.get("frequency") or norm_row.get("freq") or
            norm_row.get("usage count") or norm_row.get("usage") or "0"
        )
        try:
            freq = int(str(freq_raw).strip())
        except ValueError:
            freq = 0
        result.append({
            "source": source,
            "target": target,
            "freq": freq,
            "created": norm_row.get("created_at") or norm_row.get("date") or None,
            "created_by": norm_row.get("created_by") or norm_row.get("author") or None,
            "source_id": None,
        })
    return result


# ── Upsert ─────────────────────────────────────────────────────────────

def _upsert_rows(parsed: list, progress_callback=None) -> dict:
    now = datetime.now().isoformat()
    total_rows = len(parsed)
    invalid = sum(1 for r in parsed if r is None)
    valid = [r for r in parsed if r is not None]

    # Load existing normalized_source index once — avoid N+1 queries
    with get_connection() as conn:
        existing: dict[str, dict] = {
            row["normalized_source"]: {"id": row["id"], "frequency": row["frequency"]}
            for row in conn.execute(
                "SELECT id, normalized_source, frequency FROM translation_memory"
            ).fetchall()
        }

    to_insert: list[tuple] = []
    to_update: list[tuple] = []   # (new_freq, modified_at, id)
    seen_in_batch: set[str] = set()
    duplicates = 0

    for i, row in enumerate(valid):
        source = row["source"]
        target = row["target"]
        norm_src = normalize_segment(source)
        norm_tgt = normalize_segment(target)
        freq = max(row.get("freq") or 0, 1)
        category = detect_category(source, target)
        created = row.get("created") or now
        created_by = row.get("created_by") or "TM_IMPORT"
        source_id = row.get("source_id")

        if norm_src in existing:
            ex = existing[norm_src]
            to_update.append((ex["frequency"] + freq, now, ex["id"]))
            duplicates += 1
        elif norm_src in seen_in_batch:
            # Same segment appears multiple times in the import file
            duplicates += 1
        else:
            to_insert.append((
                source, target, norm_src, norm_tgt,
                freq, category, 1.0,
                created, now, created_by, source_id,
            ))
            seen_in_batch.add(norm_src)
            existing[norm_src] = {"id": None, "frequency": freq}

        if progress_callback and i % 500 == 0:
            progress_callback(i / len(valid))

    with get_connection() as conn:
        if to_insert:
            conn.executemany(
                "INSERT INTO translation_memory "
                "(source_segment, target_segment, normalized_source, normalized_target, "
                "frequency, category, confidence, created_at, modified_at, created_by, source_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                to_insert,
            )
        if to_update:
            conn.executemany(
                "UPDATE translation_memory SET frequency=?, modified_at=? WHERE id=?",
                to_update,
            )

    # Reload TM in-memory index so new entries are immediately available
    try:
        from engines.tm_matcher import get_matcher
        get_matcher().reload()
    except Exception:
        pass

    if progress_callback:
        progress_callback(1.0)

    return {
        "inserted": len(to_insert),
        "updated": len(to_update),
        "duplicates": duplicates,
        "invalid": invalid,
        "total": total_rows,
    }


# ── Column helpers ─────────────────────────────────────────────────────

def _find_col(header: list, candidates: list) -> int | None:
    for c in candidates:
        for i, h in enumerate(header):
            if c in h:
                return i
    return None


def _find_key(fieldnames: list, candidates: list) -> str | None:
    for c in candidates:
        for f in fieldnames:
            if c in f:
                return f
    return None


def _cell_str(row, col) -> str:
    if col is None or col >= len(row):
        return ""
    v = row[col]
    return str(v).strip() if v is not None else ""


def _cell_int(row, col) -> int:
    if col is None or col >= len(row):
        return 0
    try:
        return int(row[col]) if row[col] is not None else 0
    except (TypeError, ValueError):
        return 0
