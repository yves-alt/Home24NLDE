import re
from pathlib import Path
from dataclasses import dataclass, field

import openpyxl
import pandas as pd


@dataclass
class SheetData:
    name: str
    headers: list[str]
    source_col: int | None
    target_col: int | None
    rows: list[dict]
    df: pd.DataFrame


@dataclass
class WorkbookData:
    filename: str
    sheets: list[SheetData] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(len(s.rows) for s in self.sheets)


GERMAN_COLUMN_HINTS = [
    "de", "de-de", "deutsch", "german", "source", "original",
    "name_de", "title_de", "beschreibung", "text_de"
]
DUTCH_COLUMN_HINTS = [
    "nl", "nl-nl", "dutch", "netherlands", "target", "translation",
    "name_nl", "title_nl", "beschrijving", "text_nl"
]

GERMAN_WORD_SAMPLE = [
    "und", "oder", "mit", "für", "von", "der", "die", "das",
    "ein", "eine", "ist", "nicht", "auch",
]


def load_workbook(filepath: str, source_col_override: str | None = None, target_col_override: str | None = None) -> WorkbookData:
    path = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = WorkbookData(filename=path.name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 2:
            continue

        header_row_idx = _find_header_row(all_rows)
        headers = [str(c) if c is not None else "" for c in all_rows[header_row_idx]]

        src_col = None
        tgt_col = None

        if source_col_override:
            src_col = _col_name_to_idx(source_col_override, headers)
        if target_col_override:
            tgt_col = _col_name_to_idx(target_col_override, headers)

        if src_col is None:
            src_col = _detect_source_col(headers, all_rows[header_row_idx + 1:header_row_idx + 6])
        if tgt_col is None:
            tgt_col = _detect_target_col(headers)

        data_rows = all_rows[header_row_idx + 1:]
        parsed_rows = []
        for row in data_rows:
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            row_dict["_raw"] = row
            parsed_rows.append(row_dict)

        df = pd.DataFrame(parsed_rows)

        result.sheets.append(SheetData(
            name=sheet_name,
            headers=headers,
            source_col=src_col,
            target_col=tgt_col,
            rows=parsed_rows,
            df=df,
        ))

    wb.close()
    return result


def _find_header_row(rows: list) -> int:
    for i, row in enumerate(rows[:5]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 2:
            return i
    return 0


def _detect_source_col(headers: list[str], sample_rows: list) -> int | None:
    headers_lower = [h.lower() for h in headers]

    for hint in GERMAN_COLUMN_HINTS:
        for i, h in enumerate(headers_lower):
            if hint == h or hint in h:
                return i

    for i, h in enumerate(headers_lower):
        col_values = [str(row[i]) for row in sample_rows if i < len(row) and row[i]]
        if _looks_german(col_values):
            return i

    return 0 if headers else None


def _detect_target_col(headers: list[str]) -> int | None:
    headers_lower = [h.lower() for h in headers]
    for hint in DUTCH_COLUMN_HINTS:
        for i, h in enumerate(headers_lower):
            if hint == h or hint in h:
                return i
    return None


def _looks_german(values: list[str]) -> bool:
    if not values:
        return False
    text = " ".join(v.lower() for v in values)
    hits = sum(1 for w in GERMAN_WORD_SAMPLE if f" {w} " in f" {text} ")
    return hits >= 2


def _col_name_to_idx(name: str, headers: list[str]) -> int | None:
    name_lower = name.lower()
    for i, h in enumerate(headers):
        if h.lower() == name_lower or name_lower in h.lower():
            return i
    try:
        idx = int(name)
        if 0 <= idx < len(headers):
            return idx
    except ValueError:
        pass
    return None
