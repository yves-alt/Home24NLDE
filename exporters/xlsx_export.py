import io
import os
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TM_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FUZZY_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
AI_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
QA_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

SOURCE_TYPE_COLORS = {
    "TM_EXACT": "E8F5E9",
    "TM_FUZZY": "FFF9C4",
    "TM_PATTERN": "E3F2FD",
    "GLOSSARY": "F3E5F5",
    "CONTEXT": "E0F7FA",
    "AI": "FCE4EC",
    "EMPTY": "F5F5F5",
}


def export_workbook(
    source_filepath: str,
    sheet_results: dict,
    output_dir: str = "data/exports",
) -> str:
    source_name = Path(source_filepath).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"NL-{source_name}_{timestamp}.xlsx"
    output_path = Path(output_dir) / output_name
    output_path.parent.mkdir(parents=True, exist_ok=True)

    source_wb = openpyxl.load_workbook(source_filepath, data_only=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    for sheet_name, result_data in sheet_results.items():
        if sheet_name not in source_wb.sheetnames:
            continue

        src_ws = source_wb[sheet_name]
        src_rows = list(src_ws.iter_rows(values_only=True))

        out_ws = out_wb.create_sheet(title=sheet_name)

        headers = list(src_rows[0]) if src_rows else []
        translation_results = result_data.get("results", [])
        target_col_idx = result_data.get("target_col_idx")
        source_col_idx = result_data.get("source_col_idx")

        if target_col_idx is None:
            headers = list(headers) + ["NL Translation", "Source Type", "QA Issues"]
            target_col_idx = len(headers) - 3

        _write_header(out_ws, headers)

        result_map = {r.get("row_idx", i): r for i, r in enumerate(translation_results)}

        for row_i, row in enumerate(src_rows[1:], 1):
            row_list = list(row)

            res = result_map.get(row_i - 1)
            if res:
                if target_col_idx is not None and target_col_idx < len(row_list):
                    row_list[target_col_idx] = res.get("target", row_list[target_col_idx])
                else:
                    row_list.append(res.get("target", ""))
                    row_list.append(res.get("source_type", ""))
                    row_list.append(str(len(res.get("qa_issues", []))))

            src_type = res.get("source_type", "") if res else ""
            fill_color = SOURCE_TYPE_COLORS.get(src_type, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col_i, val in enumerate(row_list, 1):
                cell = out_ws.cell(row=row_i + 1, column=col_i, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=False, vertical="center")
                if target_col_idx is not None and col_i == target_col_idx + 1:
                    cell.fill = fill

        _auto_width(out_ws)

    _write_summary_sheet(out_wb, sheet_results)

    out_wb.save(str(output_path))
    source_wb.close()
    return str(output_path)


def export_workbook_translated_bytes(
    original_bytes: bytes,
    translation_map: dict,
    headers: list,
    sheet_name: str = "Tabelle1",
) -> bytes:
    """Return translated Excel as bytes.

    Modifies sheet_name in-place on a copy of the original workbook.
    translation_map: {row_idx (0-based data row): {col_name: dutch_value}}
    headers: ordered list of column names from row 1 of sheet_name.
    """
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    # Build header → 1-based column index map
    col_map = {h: idx + 1 for idx, h in enumerate(headers) if h}

    for row_idx, col_translations in translation_map.items():
        # row_idx is 0-based data row; Excel row = +2 (1-based + header)
        xl_row = row_idx + 2
        for col_name, dutch_value in col_translations.items():
            if col_name in col_map:
                ws.cell(row=xl_row, column=col_map[col_name], value=dutch_value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, headers: list):
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=str(h) if h else "")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 24


def _auto_width(ws, max_width: int = 60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _write_summary_sheet(wb, sheet_results: dict):
    ws = wb.create_sheet(title="Translation Summary")
    _write_header(ws, ["Sheet", "Rows", "TM Hits", "Fuzzy", "Glossary", "AI", "QA Fixes", "Consistency"])

    for i, (sheet_name, data) in enumerate(sheet_results.items(), 2):
        stats = data.get("stats", {})
        ws.cell(i, 1, sheet_name)
        ws.cell(i, 2, stats.get("total_rows", 0))
        ws.cell(i, 3, stats.get("tm_hits", 0))
        ws.cell(i, 4, stats.get("fuzzy_hits", 0))
        ws.cell(i, 5, stats.get("glossary_hits", 0))
        ws.cell(i, 6, stats.get("ai_hits", 0))
        ws.cell(i, 7, stats.get("qa_corrections", 0))
        consistency = stats.get("consistency_score", 1.0)
        ws.cell(i, 8, f"{consistency:.0%}")

    _auto_width(ws)
