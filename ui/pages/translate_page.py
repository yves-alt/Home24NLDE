import io
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

# ── Single source of truth for columns ────────────────────────────────

TRANSLATABLE_COLUMNS_NL = frozenset({
    "name",
    "colorDetail",
    "deliveryScope",
    "otherMeasurements",
    "qualityDetail",
    "textileCompositionCover1",
    "variantName",
    "materialDetail",
    "textileComposition",
})

PROTECTED_COLUMNS = frozenset({
    "articleNumber", "sku", "id", "ean", "gtin",
})

# Detection set = everything we translate + everything we protect + extra structural cols
HOME24_DETECTION_COLUMNS = TRANSLATABLE_COLUMNS_NL | PROTECTED_COLUMNS | frozenset({
    "internalDimensionDrawer", "externalDimension", "weightNetto", "weightBrutto",
    "colorName", "descriptionBullet1", "descriptionBullet2",
})


# ── Header normalization ───────────────────────────────────────────────

def _norm(h: str) -> str:
    """Trim, strip hidden chars, lowercase — for case-insensitive matching."""
    return h.strip().replace("​", "").replace("\xa0", " ").lower()


def _resolve_columns(raw_headers: list[str]) -> tuple[list[str], list[str]]:
    """Return (translatable_originals, protected_originals) via case-insensitive match.

    Returns original header names so row dicts stay compatible.
    """
    trans_norm = {_norm(c): c for c in TRANSLATABLE_COLUMNS_NL}
    prot_norm = {_norm(c): c for c in PROTECTED_COLUMNS}
    translatable, protected = [], []
    for h in raw_headers:
        n = _norm(h)
        if n in trans_norm:
            translatable.append(h)
        elif n in prot_norm:
            protected.append(h)
    return translatable, protected


# ── TranslationPlan ────────────────────────────────────────────────────

@dataclass
class TranslationPlan:
    sheet_name: str
    translatable_cols: list = field(default_factory=list)
    protected_cols: list = field(default_factory=list)
    cell_counts: dict = field(default_factory=dict)   # col -> non-empty source count

    @property
    def total_expected(self) -> int:
        return sum(self.cell_counts.values())


# ── Sheet detection ────────────────────────────────────────────────────

def _detect_best_sheet(file_bytes: bytes) -> tuple[str | None, list[tuple[str, int, list[str]]]]:
    """Score every sheet and return the best match.

    Returns (best_sheet_name or None, scored_list).
    scored_list items: (sheet_name, score, matched_column_names)
    None = ambiguous — caller must show a selector.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    scored: list[tuple[str, int, list[str]]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            scored.append((sheet_name, 0, []))
            continue
        raw = [str(c).strip() for c in first_row if c is not None and str(c).strip()]
        det_norm = {_norm(c): c for c in HOME24_DETECTION_COLUMNS}
        matched = sorted({det_norm[_norm(h)] for h in raw if _norm(h) in det_norm})
        scored.append((sheet_name, len(matched), matched))

    wb.close()
    scored.sort(key=lambda x: x[1], reverse=True)

    if len(scored) == 1:
        return scored[0][0], scored

    best_score = scored[0][1]
    if best_score == 0:
        return None, scored

    top = [s for s in scored if s[1] == best_score]
    if len(top) == 1:
        return top[0][0], scored

    return None, scored


# ── Session state ──────────────────────────────────────────────────────

_STATE_KEYS = [
    "t_step", "t_filename", "t_file_bytes", "t_preview_rows",
    "t_original_preview", "t_xl_bytes", "t_csv_bytes", "t_stats",
    "t_headers", "t_data_rows", "t_version",
    "t_detected_sheet", "t_detection_scored",
    "t_plan", "t_coverage",
]


def _init_state():
    for key in _STATE_KEYS:
        if key not in st.session_state:
            st.session_state[key] = None
    if st.session_state["t_step"] is None:
        st.session_state["t_step"] = "upload"
    if st.session_state["t_version"] is None:
        st.session_state["t_version"] = 0


def render():
    from auth.session import require_permission
    require_permission("translate")

    st.markdown('<div class="section-header">Translate Workbook</div>', unsafe_allow_html=True)
    _init_state()

    if st.session_state["t_step"] == "upload":
        _render_upload()
    else:
        _render_preview()


# ── Upload ─────────────────────────────────────────────────────────────

def _render_upload():
    from auth.credentials import get_openai_key

    if not get_openai_key():
        st.markdown(
            '<div class="alert-warning">OpenAI API key not configured. TM-only mode active.</div>',
            unsafe_allow_html=True,
        )

    uploaded = st.file_uploader(
        "Upload German Excel file",
        type=["xlsx", "xls"],
        help="The app detects the correct sheet automatically based on column structure.",
    )

    if not uploaded:
        st.markdown(
            '<div class="alert-info">Upload a German Excel file to begin. '
            "The app detects the translation sheet automatically "
            "based on Home24 column structure.</div>",
            unsafe_allow_html=True,
        )
        return

    file_bytes = uploaded.getvalue()

    try:
        detected_sheet, scored = _detect_best_sheet(file_bytes)
    except Exception as e:
        st.error(f"Cannot open file: {e}")
        return

    st.success(f"File ready: **{uploaded.name}**")

    if detected_sheet:
        _, score, matched = next((s for s in scored if s[0] == detected_sheet), (None, 0, []))
        if len(scored) > 1:
            st.info(
                f"Sheet detected: **{detected_sheet}** "
                f"— {score} Home24 column(s) matched: {', '.join(matched)}"
            )
        selected_sheet = detected_sheet
    else:
        sheet_labels = [
            f"{name} (score: {score})" if score > 0 else name
            for name, score, _ in scored
        ]
        choice_idx = st.selectbox(
            "Multiple sheets found — select the sheet to translate:",
            options=range(len(scored)),
            format_func=lambda i: sheet_labels[i],
        )
        selected_sheet = scored[choice_idx][0]

    if st.button("Translate", type="primary", use_container_width=True):
        _run_translation(file_bytes, uploaded.name, selected_sheet, scored)


# ── Translation ────────────────────────────────────────────────────────

def _run_translation(
    file_bytes: bytes,
    filename: str,
    sheet_name: str,
    scored: list[tuple[str, int, list[str]]],
):
    try:
        wb_check = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        st.error(f"Cannot open file: {e}")
        return

    if sheet_name not in wb_check.sheetnames:
        st.error(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb_check.sheetnames)}")
        wb_check.close()
        return

    ws = wb_check[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb_check.close()

    if len(all_rows) < 2:
        st.error(f"Sheet '{sheet_name}' has no data rows.")
        return

    raw_headers = [str(c) if c is not None else "" for c in all_rows[0]]
    translatable_cols, protected_cols = _resolve_columns(raw_headers)

    if not translatable_cols:
        st.error(
            f"No translatable columns found in '{sheet_name}'.\n"
            f"Expected any of: {', '.join(sorted(TRANSLATABLE_COLUMNS_NL))}\n"
            f"Found: {', '.join(h for h in raw_headers if h)}"
        )
        return

    data_rows = [
        {raw_headers[i]: row[i] for i in range(min(len(raw_headers), len(row)))}
        for row in all_rows[1:]
    ]

    # Build TranslationPlan — count non-empty source cells per column
    cell_counts = {}
    for col in translatable_cols:
        count = sum(
            1 for row in data_rows
            if row.get(col) is not None and str(row.get(col, "")).strip()
        )
        cell_counts[col] = count

    plan = TranslationPlan(
        sheet_name=sheet_name,
        translatable_cols=translatable_cols,
        protected_cols=protected_cols,
        cell_counts=cell_counts,
    )

    from engines.translation_engine import get_engine
    from engines.consistency_engine import get_consistency_engine

    engine = get_engine()
    consistency = get_consistency_engine()
    consistency.reset()
    consistency.lock_batch_from_glossary()

    preview_rows = []
    stats = {
        "tm_hits": 0, "fuzzy_hits": 0, "glossary_hits": 0,
        "phrase_hits": 0, "corpus_hits": 0, "ai_hits": 0,
        "qa_corrections": 0, "total_cells": 0,
    }

    overall_progress = st.progress(0.0)
    status_text = st.empty()
    total_cols = len(translatable_cols)

    for col_idx, col_name in enumerate(translatable_cols):
        status_text.text(
            f"Translating '{col_name}' ({col_idx + 1}/{total_cols}) "
            f"— {cell_counts.get(col_name, 0)} cells…"
        )

        items = [
            (row_idx, str(row.get(col_name, "")).strip())
            for row_idx, row in enumerate(data_rows)
            if row.get(col_name) is not None and str(row.get(col_name, "")).strip()
        ]

        if not items:
            overall_progress.progress((col_idx + 1) / total_cols)
            continue

        def _cb(p, ci=col_idx, tot=total_cols):
            overall_progress.progress((ci + p) / tot)

        batch = engine.translate_batch(
            items,
            context_rows=data_rows,
            filename=filename,
            progress_callback=_cb,
        )

        for i, (row_idx, source) in enumerate(items):
            res = batch.results[i]
            preview_rows.append({
                "Row": row_idx + 1,
                "Column": col_name,
                "German source": source,
                "Dutch translation": res.target,
                "Confidence": res.confidence_label,
                "Origin": res.source_type.value,
            })

        stats["tm_hits"] += batch.tm_hits
        stats["fuzzy_hits"] += batch.fuzzy_hits
        stats["glossary_hits"] += batch.glossary_hits
        stats["phrase_hits"] += batch.phrase_hits
        stats["corpus_hits"] += batch.corpus_hits
        stats["ai_hits"] += batch.ai_hits
        stats["qa_corrections"] += batch.qa_corrections
        stats["total_cells"] += len(items)

        overall_progress.progress((col_idx + 1) / total_cols)

    overall_progress.progress(1.0)
    status_text.text("Translation complete.")

    # Build per-column coverage stats
    coverage = _build_coverage(preview_rows, data_rows, translatable_cols)

    st.session_state["t_file_bytes"] = file_bytes
    st.session_state["t_filename"] = filename
    st.session_state["t_headers"] = raw_headers
    st.session_state["t_data_rows"] = data_rows
    st.session_state["t_preview_rows"] = preview_rows
    st.session_state["t_original_preview"] = [dict(r) for r in preview_rows]
    st.session_state["t_stats"] = stats
    st.session_state["t_xl_bytes"] = None
    st.session_state["t_csv_bytes"] = None
    st.session_state["t_version"] = 0
    st.session_state["t_step"] = "preview"
    st.session_state["t_detected_sheet"] = sheet_name
    st.session_state["t_detection_scored"] = scored
    st.session_state["t_plan"] = plan
    st.session_state["t_coverage"] = coverage

    for k in [k for k in st.session_state if k.startswith("preview_editor_")]:
        del st.session_state[k]

    st.rerun()


# ── Coverage helpers ───────────────────────────────────────────────────

def _build_coverage(
    preview_rows: list,
    data_rows: list,
    translatable_cols: list,
) -> list[dict]:
    """Build per-column coverage report rows."""
    translated_counter = Counter(r["Column"] for r in preview_rows)
    unchanged_counter = Counter(
        r["Column"] for r in preview_rows
        if r["German source"] == r["Dutch translation"]
    )

    rows = []
    for col in translatable_cols:
        source_ne = sum(
            1 for row in data_rows
            if row.get(col) is not None and str(row.get(col, "")).strip()
        )
        translated = translated_counter.get(col, 0)
        unchanged = unchanged_counter.get(col, 0)
        rows.append({
            "Column": col,
            "Source cells": source_ne,
            "Translated": translated,
            "Unchanged": unchanged,
            "Coverage": f"{translated}/{source_ne}" if source_ne else "—",
            "_ok": translated >= source_ne,
        })
    return rows


def _validate_coverage(
    preview_rows: list,
    data_rows: list,
    translatable_cols: list,
) -> list[str]:
    """Return list of error strings for any column with missing translations."""
    coverage = _build_coverage(preview_rows, data_rows, translatable_cols)
    return [
        f"'{r['Column']}': {r['Translated']} translated, {r['Source cells']} expected"
        for r in coverage
        if not r["_ok"] and r["Source cells"] > 0
    ]


# ── Preview ────────────────────────────────────────────────────────────

def _render_preview():
    from auth.session import current_user

    filename = st.session_state.get("t_filename") or "file.xlsx"
    stats = st.session_state.get("t_stats") or {}
    preview_rows = st.session_state.get("t_preview_rows") or []
    version = st.session_state.get("t_version") or 0
    detected_sheet = st.session_state.get("t_detected_sheet") or "—"
    plan: TranslationPlan | None = st.session_state.get("t_plan")
    coverage: list[dict] | None = st.session_state.get("t_coverage")

    st.markdown(f"**File:** {filename} &nbsp;·&nbsp; **Sheet:** {detected_sheet}")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("TM Exact", stats.get("tm_hits", 0))
    c2.metric("Fuzzy + Glossary", stats.get("fuzzy_hits", 0) + stats.get("glossary_hits", 0))
    c3.metric("Phrase / Corpus", stats.get("phrase_hits", 0) + stats.get("corpus_hits", 0))
    c4.metric("AI (GPT)", stats.get("ai_hits", 0))
    c5.metric("Total cells", stats.get("total_cells", 0))

    # Column coverage report
    if coverage:
        has_gap = any(not r["_ok"] and r["Source cells"] > 0 for r in coverage)
        display_cov = [{k: v for k, v in r.items() if k != "_ok"} for r in coverage]
        cov_df = pd.DataFrame(display_cov)

        if has_gap:
            st.markdown(
                '<div class="alert-error">Coverage gap detected — some columns have untranslated cells. '
                "Export is blocked until all cells are translated.</div>",
                unsafe_allow_html=True,
            )
        st.markdown("**Column coverage**")
        st.dataframe(cov_df, use_container_width=True, hide_index=True)

    # Admin debug panel
    try:
        user = current_user()
        is_admin = user and user.role == "admin"
    except Exception:
        is_admin = False

    if is_admin:
        scored = st.session_state.get("t_detection_scored") or []
        with st.expander("Detection debug", expanded=False):
            if scored:
                _, top_score, top_matched = next(
                    (s for s in scored if s[0] == detected_sheet), (None, 0, [])
                )
                st.markdown(f"**Detected sheet:** `{detected_sheet}`")
                st.markdown(f"**Detection score:** {top_score}")
                st.markdown(
                    f"**Matched columns:** {', '.join(top_matched) if top_matched else '—'}"
                )
                if len(scored) > 1:
                    st.markdown("**All sheets scored:**")
                    for name, score, matched in scored:
                        indicator = " ← selected" if name == detected_sheet else ""
                        st.markdown(
                            f"- `{name}` — score {score}"
                            + (f" ({', '.join(matched)})" if matched else "")
                            + indicator
                        )
            if plan:
                st.markdown("**Translation plan:**")
                for col in plan.translatable_cols:
                    st.markdown(f"- `{col}`: {plan.cell_counts.get(col, 0)} cells")
                st.markdown(
                    f"**Protected columns present:** "
                    + (", ".join(f"`{c}`" for c in plan.protected_cols) or "—")
                )

    st.markdown("---")
    st.markdown("### Translation preview")
    st.caption("Edit the **Dutch translation** column directly. Click **Validate** to apply edits and generate files.")

    editor_key = f"preview_editor_{version}"
    preview_df = pd.DataFrame(preview_rows) if preview_rows else pd.DataFrame(
        columns=["Row", "Column", "German source", "Dutch translation", "Confidence", "Origin"]
    )

    edited_df = st.data_editor(
        preview_df,
        column_config={
            "Row": st.column_config.NumberColumn("Row", disabled=True, width="small"),
            "Column": st.column_config.TextColumn("Column", disabled=True, width="medium"),
            "German source": st.column_config.TextColumn("German source", disabled=True, width="large"),
            "Dutch translation": st.column_config.TextColumn("Dutch translation", disabled=False, width="large"),
            "Confidence": st.column_config.TextColumn("Confidence", disabled=True, width="small"),
            "Origin": st.column_config.TextColumn("Origin", disabled=True, width="small"),
        },
        hide_index=True,
        use_container_width=True,
        height=560,
        key=editor_key,
    )

    st.markdown("---")
    col_btn, col_reset = st.columns([3, 1])

    with col_btn:
        validate_clicked = st.button(
            "Validate modifications and generate new files",
            type="primary",
            use_container_width=True,
        )

    with col_reset:
        if st.button("Start over", use_container_width=True):
            _reset_state()
            st.rerun()

    if validate_clicked:
        current_rows = edited_df.to_dict("records")
        for r in current_rows:
            r["Row"] = int(r["Row"])
        _validate_and_generate(current_rows)
        return

    # Download section — always visible once files exist
    xl_bytes = st.session_state.get("t_xl_bytes")
    csv_bytes = st.session_state.get("t_csv_bytes")

    if xl_bytes or csv_bytes:
        st.markdown("---")
        st.markdown("### Download translated files")
        stem = Path(filename).stem
        dc1, dc2 = st.columns(2)
        with dc1:
            if xl_bytes:
                st.download_button(
                    "Download NL Excel (.xlsx)",
                    data=xl_bytes,
                    file_name=f"NL-{stem}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
        with dc2:
            if csv_bytes:
                st.download_button(
                    "Download NL CSV (.csv)",
                    data=csv_bytes,
                    file_name=f"NL-{stem}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )


# ── Validate & generate ────────────────────────────────────────────────

def _validate_and_generate(current_rows: list):
    filename = st.session_state.get("t_filename") or "file.xlsx"
    file_bytes = st.session_state.get("t_file_bytes")
    original_preview = st.session_state.get("t_original_preview") or []
    headers = st.session_state.get("t_headers") or []
    data_rows = st.session_state.get("t_data_rows") or []
    sheet_name = st.session_state.get("t_detected_sheet") or "Sheet1"
    plan: TranslationPlan | None = st.session_state.get("t_plan")

    # Detect human corrections
    original_map = {(r["Row"], r["Column"]): r["Dutch translation"] for r in original_preview}
    corrections = []
    correction_lookup: dict[str, str] = {}

    for row in current_rows:
        key = (row["Row"], row["Column"])
        old_val = original_map.get(key, "")
        new_val = (row.get("Dutch translation") or "").strip()
        if new_val and new_val != old_val:
            corrections.append({
                "source": row["German source"],
                "old_target": old_val,
                "new_target": new_val,
                "column": row["Column"],
            })
            correction_lookup[row["German source"]] = new_val

    # Propagate corrections to all identical source segments
    if correction_lookup:
        for row in current_rows:
            src = row.get("German source", "")
            if src in correction_lookup:
                row["Dutch translation"] = correction_lookup[src]

    # Coverage validation — block export if any column has gaps
    translatable_cols = plan.translatable_cols if plan else list(
        {r["Column"] for r in current_rows}
    )
    coverage_errors = _validate_coverage(current_rows, data_rows, translatable_cols)
    if coverage_errors:
        for err in coverage_errors:
            st.error(f"Coverage validation failed: {err}")
        st.error("Export blocked. Fix coverage gaps before downloading.")
        return

    if corrections:
        _save_corrections(corrections, correction_lookup)

    # Build translation_map {row_idx 0-based: {col: dutch}}
    translation_map: dict[int, dict[str, str]] = {}
    for row in current_rows:
        row_idx = row["Row"] - 1
        col = row["Column"]
        translation_map.setdefault(row_idx, {})[col] = row["Dutch translation"]

    # Generate Excel bytes
    try:
        from exporters.xlsx_export import export_workbook_translated_bytes
        xl_bytes = export_workbook_translated_bytes(
            file_bytes, translation_map, headers, sheet_name=sheet_name
        )
    except Exception as e:
        st.error(f"Excel export failed: {e}")
        return

    # Generate CSV bytes
    try:
        from exporters.csv_export import generate_csv_bytes
        csv_bytes = generate_csv_bytes(headers, data_rows, translation_map)
    except Exception as e:
        st.error(f"CSV export failed: {e}")
        return

    # Update coverage stats after edits
    updated_coverage = _build_coverage(current_rows, data_rows, translatable_cols)

    st.session_state["t_preview_rows"] = current_rows
    st.session_state["t_original_preview"] = [dict(r) for r in current_rows]
    st.session_state["t_xl_bytes"] = xl_bytes
    st.session_state["t_csv_bytes"] = csv_bytes
    st.session_state["t_coverage"] = updated_coverage
    st.session_state["t_version"] = (st.session_state.get("t_version") or 0) + 1

    if corrections:
        st.success(f"Files generated. {len(corrections)} human correction(s) saved to glossary and TM.")
    else:
        st.success("Files generated successfully.")

    st.rerun()


# ── Human correction persistence ──────────────────────────────────────

def _save_corrections(corrections: list, correction_lookup: dict):
    from database.database import get_connection

    now = datetime.now().isoformat()

    for c in corrections:
        src_lower = c["source"].lower().strip()
        tgt = c["new_target"]

        try:
            with get_connection() as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO glossary "
                    "(source_term, target_term, category, confidence, source_type, active) "
                    "VALUES (?,?,?,1.0,'HUMAN_REVIEW',1)",
                    (src_lower, tgt, c["column"]),
                )
        except Exception:
            pass

        try:
            normalized_src = src_lower
            normalized_tgt = tgt.lower().strip()
            with get_connection() as conn:
                existing = conn.execute(
                    "SELECT id FROM translation_memory WHERE normalized_source=? LIMIT 1",
                    (normalized_src,),
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE translation_memory SET target_segment=?, normalized_target=?, "
                        "confidence=1.0, modified_at=?, created_by='HUMAN_REVIEW' WHERE id=?",
                        (tgt, normalized_tgt, now, existing["id"]),
                    )
                else:
                    conn.execute(
                        "INSERT INTO translation_memory "
                        "(source_segment, target_segment, normalized_source, normalized_target, "
                        "confidence, created_at, modified_at, created_by) "
                        "VALUES (?,?,?,?,1.0,?,?,'HUMAN_REVIEW')",
                        (c["source"], tgt, normalized_src, normalized_tgt, now, now),
                    )
        except Exception:
            pass

    try:
        from engines.translation_engine import get_engine
        eng = get_engine()
        for src, tgt in correction_lookup.items():
            eng._dedup_cache[src] = tgt
    except Exception:
        pass


# ── Reset ──────────────────────────────────────────────────────────────

def _reset_state():
    for key in _STATE_KEYS:
        st.session_state[key] = None
    st.session_state["t_step"] = "upload"
    st.session_state["t_version"] = 0
    for k in [k for k in st.session_state if k.startswith("preview_editor_")]:
        del st.session_state[k]
